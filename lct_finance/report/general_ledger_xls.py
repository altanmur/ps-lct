# -*- encoding: utf-8 -*-
##############################################################################
#
#    Odoo, Open Source Management Solution
#    Copyright (C) 2004-TODAY Odoo S.A. <http://www.openerp.com>
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as
#    published by the Free Software Foundation, either version 3 of the
#    License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
##############################################################################


import xlwt
from datetime import datetime
from openerp.addons.report_xls.report_xls import report_xls
from openerp.addons.report_xls.utils import rowcol_to_cell, _render
# from .general_ledger import general_ledger_report
from openerp.addons.account.report.account_general_ledger import general_ledger
from openerp.tools.translate import _
import logging
_logger = logging.getLogger(__name__)

# Some nonsense because I'm monkey patching...
from .general_ledger import patch_lines, patch_get_children_accounts
general_ledger.lines = patch_lines
general_ledger.get_children_accounts = patch_get_children_accounts


class general_ledger_xls_parser(general_ledger):

    def __init__(self, cr, uid, name, context):
        super(general_ledger_xls_parser, self).__init__(cr, uid, name, context=context)
        self.context = context
        wanted_list = self._report_xls_fields(cr, uid, context)
        template_changes = self._report_xls_template(cr, uid, context)
        space_extra = self._report_xls_render_space_extra(cr, uid, context)
        self.localcontext.update({
            'datetime': datetime,
            'wanted_list': wanted_list,
            'template_changes': template_changes,
            'space_extra': space_extra,
        })

    # Most of these do nothing, but the report kind of requires them.

    # override list in inherited module to add/drop columns or change order
    def _report_xls_fields(self, cr, uid, context=None):
        res = [
            'ldate',
            'acode',
            'period_code',
            'lcode',
            'partner_name',
            'move',
            'lname',
            'line_corresp',
            'debit',
            'credit',
            'progress',
        ]
        return res

    # allow inherited modules to extend the query
    def _report_xls_query_extra(self, cr, uid, context=None):
        select_extra = ""
        join_extra = ""
        where_extra = ""
        return (select_extra, join_extra, where_extra)

    # allow inherited modules to add document references
    def _report_xls_document_extra(self, cr, uid, context=None):
        return "''"

    # allow inherited modules to extend the render namespace
    def _report_xls_render_space_extra(self, cr, uid, context=None):
        return None

    # Change/Add Template entries
    def _report_xls_template(self, cr, uid, context=None):
        return {}

class general_ledger_xls(report_xls):

    def __init__(self, name, table, rml=False, parser=False, header=True, store=False):
        super(general_ledger_xls, self).__init__(name, table, rml, parser, header, store)

        # Cell Styles
        _xs = self.xls_styles
        # header
        rh_cell_format = _xs['bold'] + _xs['fill'] + _xs['borders_all']
        self.rh_cell_style = xlwt.easyxf(rh_cell_format)
        self.rh_cell_style_center = xlwt.easyxf(rh_cell_format + _xs['center'])
        self.rh_cell_style_right = xlwt.easyxf(rh_cell_format + _xs['right'])
        # lines
        acc_cell_format = _xs['borders_all']
        self.acc_cell_style = xlwt.easyxf(acc_cell_format)
        self.acc_cell_style_center = xlwt.easyxf(acc_cell_format + _xs['center'])
        self.acc_cell_style_date = xlwt.easyxf(acc_cell_format + _xs['left'], num_format_str=report_xls.date_format)
        self.acc_cell_style_decimal = xlwt.easyxf(acc_cell_format + _xs['right'], num_format_str=report_xls.decimal_format)
        # totals
        rt_cell_format = _xs['bold'] + _xs['fill'] + _xs['borders_all']
        self.rt_cell_style = xlwt.easyxf(rt_cell_format)
        self.rt_cell_style_right = xlwt.easyxf(rt_cell_format + _xs['right'])
        self.rt_cell_style_decimal = xlwt.easyxf(rt_cell_format + _xs['right'], num_format_str=report_xls.decimal_format)

        # XLS Template
        self.col_specs_lines_template = {
            'ldate': {
                'header': [1, 20, 'text', 'Date'],
                'lines': [1, 0, 'text', _render("l['ldate']")],
                'totals': [1, 0, 'text', None],
            },
            'acode': {
                'header': [1, 20, 'text', 'Code'],
                'lines': [1, 0, 'text', _render("l['acode']")],
                'totals': [1, 0, 'text', None],
            },
            'period_code': {
                'header': [1, 20, 'text', 'Period'],
                'lines': [1, 0, 'text', _render("l['period_code']")],
                'totals': [1, 0, 'text', None],
            },
            'lcode': {
                'header': [1, 20, 'text', 'JRNL', None, self.rh_cell_style_center],
                'lines': [1, 0, 'text', _render("l['lcode']")],
                'totals': [1, 0, 'text', None],
            },
            'partner_name': {
                'header': [1, 20, 'text', 'Partner', None, self.rh_cell_style_center],
                'lines': [1, 0, 'text', _render("l['partner_name']")],
                'totals': [1, 0, 'text', None],
            },
            'move': {
                'header': [1, 20, 'text', 'Move', None, self.rh_cell_style_center],
                'lines': [1, 0, 'text', _render("l['move']")],
                'totals': [1, 0, 'text', None],
            },
            'lname': {
                'header': [1, 20, 'text', 'Entry Label', None, self.rh_cell_style_center],
                'lines': [1, 0, 'text', _render("l['lname']")],
                'totals': [1, 0, 'text',None],
            },
            'line_corresp': {
                'header': [1, 20, 'text', 'Counterpart', None, self.rh_cell_style_center],
                'lines': [1, 0, 'text', _render("l['line_corresp']")],
                'totals': [1, 0, 'text',None],
            },
            'debit': {
                'header': [1, 20, 'text', 'Debit', None, self.rh_cell_style_center],
                'lines': [1, 0, 'number', _render("l['debit']")],
                'totals': [1, 0, 'number', None, _render("debit_formula")],
            },
            'credit': {
                'header': [1, 20, 'text', 'Credit', None, self.rh_cell_style_center],
                'lines': [1, 0, 'number', _render("l['credit']")],
                'totals': [1, 0, 'number', None, _render("credit_formula")],
            },
            'progress': {
                'header': [1, 20, 'text', 'Solde', None, self.rh_cell_style_center],
                'lines': [1, 0, 'number', _render("l['progress']")],
                'totals': [1, 0, 'number', None, _render("bal_formula")],
            },
        }

    def _account_title(self, o, ws, parser, row_pos, xlwt, _xs):
        cell_style = xlwt.easyxf(_xs['xls_title'])
        cell_style_center = xlwt.easyxf(_xs['xls_title'] + _xs['center'])
        row_specs = [
            [
                ('title', 10 + self.padding, 0, 'text', 'General Ledger', None, cell_style_center),
            ],
        ]
        for c_specs in row_specs:
            row_data = self.xls_row_template(c_specs, [x[0] for x in c_specs])
            row_pos = self.xls_write_row(ws, row_pos, row_data, row_style=cell_style)
        return row_pos + 1

    def _meta_headers(self, o, ws, parser, row_pos, xlwt, _xs):
        cell_style_center = xlwt.easyxf(_xs['bold'] + _xs['center'])
        data = parser['data']
        disp_acc = {
            'all': 'All',
            'movement': 'With movements',
            'not_zero': 'With balance is not equal to 0'}[data['form']['display_account']]
        filter_by = data['form']['filter']
        if filter_by == 'filter_no':
            filter_header = [
                ('filter_by', 2, 0, 'text', parser[('get_filter')](data)),
            ]
        elif filter_by == 'filter_date':
            filter_header = [
                ('start', 1, 0, 'text', 'Start Date'),
                ('end', 1, 0, 'text', 'End Date'),
            ]
            filter_val = [
                ('start', 1, 0, 'text', parser['get_start_date'](data)),
                ('end', 1, 0, 'text', parser['get_end_date'](data)),
            ]
        elif filter_by == 'filter_period':
            filter_header = [
                ('start', 1, 0, 'text', 'Start Period'),
                ('end', 1, 0, 'text', 'End Period'),
            ]
            filter_val = [
                ('filter_by', 1, 0, 'text', parser['get_start_period'](data)),
                ('end', 1, 0, 'text', parser['get_end_period'](data)),
            ]

        row_specs = [
            [
                ('chart_of_acc', 2 + self.padding, 0, 'text', 'Chart of Accounts', None, cell_style_center),
                ('fisc_yr', 1, 0, 'text', 'Fiscal Year', None, cell_style_center),
                ('jrnl', 2, 0, 'text', 'Journals', None, cell_style_center),
                ('disp_acc', 1, 0, 'text', 'Display Account', None, cell_style_center),
                ('filter_by', 2, 0, 'text', 'Filter By', None, cell_style_center),
                ('sorted_by', 1, 0, 'text', 'Entries Sorted By', None, cell_style_center),
                ('target_moves', 1, 0, 'text', 'Target Moves', None, cell_style_center),
            ],
            [
                ('chart_of_acc', 2 + self.padding, 0, 'text', o.name),
                ('fisc_yr', 1, 0, 'text', parser['get_fiscalyear'](data)),
                ('jrnl', 2, 0, 'text', ', '.join([ lt or '' for lt in parser['get_journal'](data) ])),
                ('disp_acc', 1, 0, 'text', disp_acc),
            ] + filter_header + [
                ('sorted_by', 1, 0, 'text', parser['get_sortby'](data)),
                ('target_moves', 1, 0, 'text', parser['get_target_move'](data)),
            ],
        ]
        if filter_by != 'filter_no':
            row_specs.append(
                [
                    ('chart_of_acc', 2 + self.padding, 0, 'text', ''),
                    ('fisc_yr', 1, 0, 'text', ''),
                    ('jrnl', 2, 0, 'text', ''),
                    ('disp_acc', 1, 0, 'text', ''),
                ] + filter_val + [
                    ('sorted_by', 1, 0, 'text', ''),
                    ('target_moves', 1, 0, 'text', ''),
                ])
        for c_specs in row_specs:
            row_data = self.xls_row_template(c_specs, [x[0] for x in c_specs])
            row_pos = self.xls_write_row(ws, row_pos, row_data)
        return row_pos

    def _print_acc_header(self, account, ws, row_pos, _xs, debit, credit, balance):
        style = xlwt.easyxf(_xs['bold'])
        header_spec = [
            ('none', 1, 0, 'text', None, None, style),
            ('cod', 1, 0, 'text', "%s %s" % (account.code, account.name), None, style),
            ('none2', 5 + self.padding, 0, 'text', None, None, style),
            ('debit', 1, 0, 'number', debit, None, style),
            ('credit', 1, 0, 'number', credit, None, style),
            ('balance', 1, 0, 'number', balance, None, style),
        ]
        row_data = self.xls_row_template(header_spec, [x[0] for x in header_spec])
        row_pos = self.xls_write_row(ws, row_pos, row_data)
        return row_pos

    def _account_lines(self, o, ws, parser, row_pos, xlwt, _xs):
        """
            o: account.account browse record
            ws: worksheet
            parser: parser
            row_pos: int
            xlwt: xlwt module
            _xs: xls styles
        """

        if parser.space_extra:
            locals().update(parser.space_extra)
        wanted_list = self.wanted_list
        debit_pos = self.debit_pos
        credit_pos = self.credit_pos


        # Column headers
        c_specs = map(lambda x: self.render(x, self.col_specs_lines_template, 'header', render_space={'_': _}), wanted_list)
        row_data = self.xls_row_template(c_specs, [x[0] for x in c_specs])
        row_pos = self.xls_write_row(ws, row_pos, row_data, row_style=self.rh_cell_style, set_column_size=True)
        ws.set_horz_split_pos(row_pos)

        # check # lines for totals formula position
        for child_acc in parser['get_children_accounts'](o):
            acc_cnt = len(parser.lines(child_acc))
            header_pos = row_pos
            row_pos += 1  # Make room for the account header
            acc_start_pos = row_pos
            # Totals on top
            debit = parser['sum_debit_account'](child_acc)
            credit = parser['sum_credit_account'](child_acc)
            balance = parser['sum_balance_account'](child_acc)
            # Account header with totals
            row_pos = self._print_acc_header(child_acc, ws, header_pos, _xs, debit, credit, balance)
            cnt = 0
            # for l in parser.lines(o):
            for l in parser.lines(child_acc):
                l['ldate'] = l['ldate'].split()[0]
                l['acode'] = child_acc.code
                cnt += 1
                debit_cell = rowcol_to_cell(row_pos, debit_pos)
                credit_cell = rowcol_to_cell(row_pos, credit_pos)
                bal_formula = debit_cell + '-' + credit_cell
                c_specs = map(lambda x: self.render(x, self.col_specs_lines_template, 'lines'), wanted_list)
                row_data = self.xls_row_template(c_specs, [x[0] for x in c_specs])
                row_pos = self.xls_write_row(ws, row_pos, row_data, row_style=self.acc_cell_style)
                if 'draw_line' in l and l['draw_line'] and cnt != acc_cnt:
                    row_pos += 1
        return row_pos + 1

    def generate_xls_report(self, parser, _xs, data, objects, wb):
        """
            parser: parser
            _xs: xls Styles
            data: {}
            objects: account.account browse objects
            wb: xls workbook
        """

        if not parser['data']['form']['amount_currency']:
            parser.wanted_list.remove('line_corresp')
            self.padding = 0
        else:
            self.padding = 1
        self.wanted_list = wanted_list = parser.wanted_list
        self.col_specs_lines_template.update(parser.template_changes)

        self.debit_pos = 'debit' in wanted_list and wanted_list.index('debit')
        self.credit_pos = 'credit' in wanted_list and wanted_list.index('credit')
        # if not (self.credit_pos and self.debit_pos) and 'balance' in wanted_list:
        #     raise orm.except_orm(_('Customisation Error!'),
        #         _("The 'Balance' field is a calculated XLS field requiring the presence of the 'Debit' and 'Credit' fields !"))

        for o in objects:
            sheet_name = 'Balance des comptes'
            ws = wb.add_sheet(sheet_name)
            ws.panes_frozen = True
            ws.remove_splits = True
            ws.portrait = 0  # Landscape
            ws.fit_width_to_pages = 1
            row_pos = 0

            # set print header/footer
            ws.header_str = self.xls_headers['standard']
            ws.footer_str = self.xls_footers['standard']

            # Top
            row_pos = self._account_title(o, ws, parser, row_pos, xlwt, _xs)
            # Near top
            row_pos = self._meta_headers(o, ws, parser, row_pos, xlwt, _xs)
            # Data
            row_pos = self._account_lines(o, ws, parser, row_pos, xlwt, _xs)

general_ledger_xls('report.xls.general_ledger', 'account.account',
    parser=general_ledger_xls_parser)


from openpyxl import Workbook
from openpyxl.styles import Style, Font, Alignment, Border, Side, colors
import cStringIO

from openerp import pooler
from openerp.tools import DEFAULT_SERVER_DATETIME_FORMAT
from openerp.report.report_sxw import report_sxw

COLUMN_WIDTH = 15

class general_ledger_xlsx_parser(general_ledger):

    def __init__(self, cr, uid, name, context):
        super(general_ledger_xlsx_parser, self).__init__(cr, uid, name, context=context)
        self.context = context
        self.localcontext.update({
            'datetime': datetime,
        })


class general_ledger_xlsx(report_sxw):

    def __init__(self, name, table, rml=False, parser=False, header=True, store=False):
        super(general_ledger_xlsx, self).__init__(name, table, rml, parser, header, store)
        alignment_center = Alignment(horizontal='center')
        self.styles = {
            'normal': Style(font=Font(name='Arial', size=10)),
            'title': Style(font=Font(bold=True, name='Arial', size=12), alignment=alignment_center),
            'label': Style(font=Font(bold=True, name='Arial', size=10), alignment=alignment_center),
            'label2': Style(font=Font(bold=True, name='Arial', size=10), alignment=alignment_center, border=Border(bottom=Side(border_style='thin', color=colors.BLACK, ))),
        }

        self.style_classes = {
            'font': Font,
            'alignment': Alignment,
            'border': Border
        }

    def _get_style(self, name=None):
        if name not in self.styles:
            name = 'normal'
        return self.styles[name]

    def _set_cell(self, ws, pos1, pos2=None, value=None, style=None):
        if pos2 is not None:
            ws.merge_cells(None, *(pos1 + pos2))

        cell = ws.cell(None, *pos1)
        cell.value = value
        cell.style = self._get_style(style)

    def create(self, cr, uid, ids, data, context=None):
        self.pool = pooler.get_pool(cr.dbname)
        self.cr = cr
        self.uid = uid
        report_obj = self.pool.get('ir.actions.report.xml')
        report_ids = report_obj.search(
            cr, uid, [('report_name', '=', self.name[7:])], context=context)
        self.table = data.get('model') or self.table
        return self.create_source_xlsx(cr, uid, ids, data, context)

    def create_source_xlsx(self, cr, uid, ids, data, context=None):
        if not context:
            context = {}
        parser_instance = self.parser(cr, uid, self.name2, context)
        self.parser_instance = parser_instance
        self.context = context
        objs = self.getObjects(cr, uid, ids, context)
        parser_instance.set_context(objs, data, ids, 'xls')
        objs = parser_instance.localcontext['objects']

        wb = Workbook()
        local_context = parser_instance.localcontext

        self.generate_xlsx_report(parser_instance.localcontext, data, objs, wb)

        n = cStringIO.StringIO()
        wb.save(n)
        n.seek(0)
        return (n.read(), 'xlsx')

    def _write_header(self, parser, data, objects, ws):
        self._set_cell(ws, [1, 1], [1, 9], "General Ledger", style='title')

        display_account = {
            'all': 'All',
            'movement': 'With movements',
            'not_zero': 'With balance is not equal to 0'
        }[data['form']['display_account']]

        filter_by = data['form']['filter']
        if filter_by == 'filter_date':
            filter_cells = [
                ([(4, 7)], 'Start Date'),
                ([(4, 8)], 'End Date'),
                ([(5, 7)], parser['get_start_date'](data)),
                ([(5, 8)], parser['get_end_date'](data)),
            ]
        elif filter_by == 'filter_period':
            filter_cells = [
                ([(4, 7)], 'Start Period'),
                ([(4, 8)], 'End Period'),
                ([(5, 7)], parser['get_start_period'](data)),
                ([(5, 8)], parser['get_end_period'](data)),
            ]
        else:
            filter_cells = [
                ([(4, 7), (4, 8)], "No filters"),
            ]

        cells_by_style = {
            'label': [
                ([(3, 1), (3, 2)], "Chart of Accounts"),
                ([(3, 3)], "Fiscal Year"),
                ([(3, 4), (3, 5)], "Journals"),
                ([(3, 6)], "Display Account"),
                ([(3, 7), (3, 8)], "Filter By"),
                ([(3, 9)], "Entries Sorted By"),
                ([(3, 10)], "Target Moves"),
            ],
            'normal': [
                ([(4, 1), (4, 2)], objects and objects[0].name),
                ([(4, 3)], parser['get_fiscalyear'](data)),
                ([(4, 4), (4, 5)], ', '.join([ lt or '' for lt in parser['get_journal'](data) ])),
                ([(4, 6)], display_account),
                ([(4, 9)], parser['get_sortby'](data)),
                ([(4, 10)], parser['get_target_move'](data)),
            ] + filter_cells,
            'label2': [
                ([(6, 1)], "Date"),
                ([(6, 2)], "Code"),
                ([(6, 3)], "Period"),
                ([(6, 4)], "JRNL"),
                ([(6, 5)], "Partner"),
                ([(6, 6)], "Move"),
                ([(6, 7)], "Entry Label"),
                ([(6, 8)], "Debit"),
                ([(6, 9)], "Credit"),
                ([(6, 10)], "Solde"),
            ]
        }

        for style, cells in cells_by_style.iteritems():
            for cell in cells:
                self._set_cell(ws, *cell[0], value=cell[1], style=style)

    def _format_sheet(self, ws):
        for i in xrange(10):
            ws.cell(row=1, column=i+1)
        for dimension in ws.column_dimensions.itervalues():
            dimension.width = COLUMN_WIDTH

    def generate_xlsx_report(self, parser, data, objects, wb):
        """
            parser: parser
            data: {}
            objects: account.account browse objects
            wb: xlsx workbook
        """

        ws = wb.active

        self._format_sheet(ws)

        self._write_header(parser, data, objects, ws)


general_ledger_xlsx('report.xlsx.general_ledger', 'account.account',
    parser=general_ledger_xlsx_parser)
